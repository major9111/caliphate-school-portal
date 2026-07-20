"""Excel/CSV export service using openpyxl."""
import io
from datetime import datetime, timezone
from typing import List, Dict, Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


PRIMARY_HEX = "1e40af"
HEADER_FILL = PatternFill("solid", fgColor=PRIMARY_HEX)
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
ALT_FILL    = PatternFill("solid", fgColor="eff6ff")
BORDER_SIDE = Side(style="thin", color="d1d5db")
THIN_BORDER = Border(left=BORDER_SIDE, right=BORDER_SIDE, top=BORDER_SIDE, bottom=BORDER_SIDE)


def _style_header_row(ws, row: int, num_cols: int):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _style_data_row(ws, row: int, num_cols: int, alternate: bool):
    fill = ALT_FILL if alternate else PatternFill()
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        if alternate:
            cell.fill = fill
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="center")


def generate_excel(title: str, headers: List[str], rows: List[List[Any]],
                   sheet_name: str = "Data") -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    # Title row
    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    title_cell = ws["A1"]
    title_cell.value = title
    title_cell.font = Font(bold=True, size=13, color=PRIMARY_HEX)
    title_cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 24

    # Generated timestamp
    ws.merge_cells(f"A2:{get_column_letter(len(headers))}2")
    ts_cell = ws["A2"]
    ts_cell.value = f"Generated: {datetime.now(timezone.utc).strftime('%d %b %Y %H:%M')} UTC"
    ts_cell.font = Font(italic=True, size=9, color="6b7280")
    ts_cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 16

    # Headers (row 3)
    for col_i, header in enumerate(headers, 1):
        ws.cell(row=3, column=col_i, value=header)
    _style_header_row(ws, 3, len(headers))
    ws.row_dimensions[3].height = 20

    # Data rows
    for row_i, row in enumerate(rows, 4):
        for col_i, val in enumerate(row, 1):
            ws.cell(row=row_i, column=col_i, value=val)
        _style_data_row(ws, row_i, len(headers), (row_i - 4) % 2 == 1)
        ws.row_dimensions[row_i].height = 18

    # Auto-fit column widths
    for col_i, header in enumerate(headers, 1):
        max_len = len(str(header))
        for row in rows:
            if col_i - 1 < len(row):
                max_len = max(max_len, len(str(row[col_i - 1])))
        ws.column_dimensions[get_column_letter(col_i)].width = min(max_len + 4, 40)

    # Freeze header row
    ws.freeze_panes = "A4"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_results_excel(results: List[Dict[str, Any]], term: str, session: str) -> bytes:
    headers = ["#", "Student Name", "Class", "Subject", "CA (40)", "Exam (60)", "Total (100)", "Grade", "Remark"]
    rows = []
    for i, r in enumerate(results, 1):
        rows.append([
            i, r.get("student_name",""), r.get("class_name",""), r.get("subject",""),
            r.get("ca_score",""), r.get("exam_score",""), r.get("total",""),
            r.get("grade",""), r.get("remark",""),
        ])
    return generate_excel(f"Examination Results — {term} {session}", headers, rows, "Results")


def generate_students_excel(students: List[Dict[str, Any]]) -> bytes:
    headers = ["#", "Admission No", "First Name", "Last Name", "Class", "Gender", "Email", "Phone", "Status"]
    rows = []
    for i, s in enumerate(students, 1):
        rows.append([
            i, s.get("admission_number",""), s.get("first_name",""), s.get("last_name",""),
            s.get("class_name",""), s.get("gender",""), s.get("email",""),
            s.get("phone",""), s.get("enrollment_status",""),
        ])
    return generate_excel("Student List", headers, rows, "Students")


def generate_payments_excel(payments: List[Dict[str, Any]]) -> bytes:
    headers = ["#", "Receipt No", "Student Name", "Type", "Method", "Amount (₦)", "Date", "Status"]
    rows = []
    for i, p in enumerate(payments, 1):
        rows.append([
            i, p.get("receipt_number",""), p.get("student_name",""), p.get("type",""),
            p.get("method",""), p.get("amount",0), p.get("payment_date",""), p.get("status",""),
        ])
    return generate_excel("Payment Records", headers, rows, "Payments")


def generate_attendance_excel(records: List[Dict[str, Any]], date: str = "") -> bytes:
    headers = ["#", "Student Name", "Class", "Date", "Status", "Marked At"]
    rows = []
    for i, r in enumerate(records, 1):
        rows.append([
            i, r.get("student_name",""), r.get("class_name",""),
            r.get("date",""), r.get("status","").title(), r.get("marked_at","")
        ])
    return generate_excel(f"Attendance Records{f' — {date}' if date else ''}", headers, rows, "Attendance")


def generate_payroll_excel(payroll_entries: List[Dict[str, Any]]) -> bytes:
    headers = ["#","Employee","ID","Role","Basic (₦)","Allowances (₦)","Deductions (₦)","Net Pay (₦)","Month","Year","Status"]
    rows = []
    total_net = 0
    for i, p in enumerate(payroll_entries, 1):
        net = float(p.get("net_salary", 0))
        total_net += net
        rows.append([
            i, p.get("employee_name",""), p.get("employee_id",""), p.get("role",""),
            p.get("basic_salary",0), p.get("allowances",0), p.get("deductions",0),
            net, p.get("month",""), p.get("year",""), p.get("status",""),
        ])
    rows.append(["","TOTAL","","","","","",total_net,"","",""])
    return generate_excel("Payroll Report", headers, rows, "Payroll")
